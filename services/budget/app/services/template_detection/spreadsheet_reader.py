"""Spreadsheet structure detection utilities.

Public API
- ExcelStructureDetector(file_path)
    - read_sheet_with_pandas() -> DataFrame
    - read_sheet_with_openpyxl() -> (DataFrame, formula_flags)
    - filter_out_formula_rows(df) -> DataFrame
    - filter_out_total_rows(df) -> DataFrame
    - normalize_dataframe(df) -> DataFrame
    - remove_numeric_rows(df) -> DataFrame
    - to_detection_json(df) -> str
    - detect_structure() -> DataFrame  # high-level pipeline, values preserved
- to_extraction_grid(df) -> list[list[str | None]]

These helpers read Excel sheets and return a cleaned DataFrame where
formula cells and total/grand-total rows are removed, ready for downstream
AI extraction. to_extraction_grid() turns that DataFrame into the plain
grid sent to the AI extraction call and hashed for the structure fingerprint.
"""

from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# from ..normalizer import normalize_label
import re
import pandas as pd

# numeric_pattern = re.compile(r"^\s*[-+]?\d*\.?\d+\s*$")
numeric_pattern = re.compile(
    r"""
    ^\s*
    (
        [-+]?\d+(\.\d*)?      # allows 1, 1., 1.5, 2.0
        |
        [-+]?\d+\s*-\s*\d+    # numeric range
    )
    \s*$
    """,
    re.VERBOSE,
)


def is_numeric(val):
    """Return True if value is numeric-like or empty/None.

    This is a permissive check used to filter out rows that are
    primarily numeric (e.g. amounts) when detecting textual labels.
    """

    if val is None:
        return True
    val_str = str(val).strip()
    if val_str == "":
        return True
    return bool(numeric_pattern.match(val_str))


def classify_row_label(text: str) -> str | None:
    """Classify a row's leading label as a total/grand-total marker, if any.

    Ported from the retired detector.py's classify_row: a category-total row
    starts with "total " but isn't the project grand total; "total project"
    (or equivalent) marks the grand total. Anything else returns None.
    """
    t = text.lower().strip()
    if not t:
        return None
    if t.startswith("total ") and "project" not in t:
        return "category_total"
    if "total project" in t:
        return "grand_total"
    return None


class ExcelStructureDetector:
    def __init__(self, file_path):
        """`file_path` may be a filesystem path (str) or a file-like object
        (e.g. io.BytesIO of an uploaded file's bytes) — openpyxl accepts
        either. Callers handling an upload should pass the in-memory bytes
        directly rather than writing to local disk first."""

        self.file_path = file_path
        self.data: list[list[Any]] = []
        self.cached_data: list[list[Any]] = []
        self.formula_flags: list[list[Any]] = []
        self.wb = load_workbook(file_path, data_only=False)
        self.ws = self.wb.active
        if hasattr(file_path, "seek"):
            file_path.seek(0)
        # Second load with cached (last-calculated) values — a formula cell's
        # `.value` under data_only=False is the formula string itself (e.g.
        # "=E18*450"), not a usable number. Real donor templates sometimes
        # compute one currency column from another this way (see
        # filter_out_formula_rows); this parallel workbook is how that
        # computed value survives into the extraction grid instead of being
        # blanked.
        self.wb_cached = load_workbook(file_path, data_only=True)
        self.ws_cached = self.wb_cached.active

    def read_sheet_with_pandas(self) -> pd.DataFrame:
        """Read the active sheet into a pandas DataFrame (text only).

        Keeps everything as strings and does not treat empty cells as NaN.
        """
        return pd.read_excel(
            self.file_path,
            header=None,  # ❗ critical
            dtype=str,  # keep everything as text
            keep_default_na=False,
        )

    def read_sheet_with_openpyxl(self) -> tuple[pd.DataFrame, list[list[Any]]]:
        """Read the sheet using openpyxl and also return formula flags.

        Returns a tuple (DataFrame, formula_flags) where formula_flags is
        a list of rows containing booleans and integers for formula tracking.
        """
        first_row = [""] + [get_column_letter(i + 1) for i in range(len(self.ws[1]))]
        self.data = [first_row]
        self.cached_data = [first_row]
        self.formula_flags = [first_row]
        rows = zip(
            self.ws.iter_rows(values_only=False), self.ws_cached.iter_rows(values_only=False)
        )
        for i, (row, cached_row) in enumerate(rows):
            row_values = [i + 1]
            row_cached = [i + 1]
            row_formulas = [i + 1]
            for cell, cached_cell in zip(row, cached_row):
                is_formula = cell.data_type == "f"  # 'f' indicates formula in openpyxl
                row_values.append(cell.value)
                row_cached.append(cached_cell.value)
                row_formulas.append(is_formula)
            self.data.append(row_values)
            self.cached_data.append(row_cached)
            self.formula_flags.append(row_formulas)

        return pd.DataFrame(self.data), self.formula_flags

    def filter_out_formula_rows(self) -> pd.DataFrame:
        """
        Resolve individual formula cells to their last-calculated cached
        value, rather than dropping their whole row or blanking them.

        A realistic template (e.g. one with a computed second-currency
        column, such as a local-currency amount derived as
        `=donor_currency_col*rate`) can have a formula in one cell of an
        otherwise-normal line-item row. Dropping the whole row on any
        formula cell discards real data; blanking just the formula cell
        (the previous fix) avoids that but throws away a value that may be
        the only place a needed figure (e.g. the local-currency amount)
        appears. Falls back to None only when the workbook has no cached
        result for a formula cell — i.e. it was never calculated by a
        spreadsheet application before upload.
        """
        # Convert formula_flags to DataFrame for easy cell-wise operations
        ff_df = pd.DataFrame(self.formula_flags)
        df = pd.DataFrame(self.data)
        cached_df = pd.DataFrame(self.cached_data)
        # We assume first column is Excel row numbers, so skip it
        ff_df_body = ff_df.iloc[1:, 1:].astype(bool)

        # Replace formula cells with their cached calculated value.
        df.iloc[1:, 1:] = df.iloc[1:, 1:].mask(ff_df_body, cached_df.iloc[1:, 1:])

        return df

    def filter_out_total_rows(self, df: pd.DataFrame) -> pd.DataFrame:
        """Drop category-total and grand-total rows so they aren't line items."""
        header = df.iloc[0:1]
        body = df.iloc[1:]
        first_col = body.columns[1]

        is_total = body[first_col].apply(
            lambda v: isinstance(v, str) and classify_row_label(v) is not None
        )
        body = body[~is_total]

        return pd.concat([header, body]).reset_index(drop=True)

    def normalize_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Remove empty rows/columns except header/row number column"""
        header = df.iloc[0]
        body = df.iloc[1:]

        # drop completely empty rows (ignore ExcelRow column)
        body = body.dropna(how="all", subset=body.columns[1:])
        return pd.concat([header.to_frame().T, body]).reset_index(drop=True)
        # return df.replace(r"^\s*$", None, regex=True).dropna(how="all").reset_index(drop=True)

    def remove_numeric_rows(self, df: pd.DataFrame) -> pd.DataFrame:
        """Remove rows that are entirely numeric (except first column and row)."""

        # first row is header (A, B, C...) and first column is row numbers / numeration
        header = df.iloc[0:1]
        body = df.iloc[1:]

        # Create a boolean mask: True for cells that are numeric-like
        mask_numeric = body.iloc[:, 1:].applymap(
            lambda v: bool(v) and bool(numeric_pattern.match(str(v))) if v is not None else True
        )

        # Replace numeric-like values with None
        body.iloc[:, 1:] = body.iloc[:, 1:].mask(mask_numeric, other=None)

        # Drop rows that are all NaN except the first column
        body_cleaned = body.dropna(how="all", subset=body.columns[1:]).reset_index(drop=True)

        # Combine with header
        df_cleaned = pd.concat([header, body_cleaned]).reset_index(drop=True)
        return df_cleaned

    def filter_list_of_possible_fields(self, df: pd.DataFrame) -> list[str]:
        """Extract a list of possible field names from the cleaned DataFrame."""
        possible_fields = set()

        for r_idx, row in df.iloc[1:].iterrows():  # skip header
            for c_idx, val in enumerate(row[1:], start=1):  # skip first column (row numbers)
                if val and isinstance(val, str) and not is_numeric(val):
                    possible_fields.add(val.strip())

        return list(possible_fields)

    def to_detection_json(self, df: pd.DataFrame) -> list[dict]:
        """Serialize detection results (from cleaned DataFrame) to a JSON string."""

        output = []

        for r_idx, row in df.iloc[1:].iterrows():  # skip header
            for c_idx, val in enumerate(row[1:], start=1):  # skip first column (row numbers)
                if not is_numeric(val):
                    coordinate = f"{df.iloc[0, c_idx]}{row[0]}"  # column letter + row number
                    # For now, dummy suggested field + confidence
                    suggested_field = "unknown"
                    confidence = 0.0
                    output.append(
                        {
                            "coordinate": coordinate,
                            "row": row[0],
                            "col": df.iloc[0, c_idx],
                            "value": val,
                            "suggested_field": suggested_field,
                            "confidence": confidence,
                        }
                    )

        # Dump to JSON
        # json_str = json.dumps(output, indent=2)
        return output

    def detect_structure(self) -> pd.DataFrame:
        """High-level pipeline: read, blank formula cells, drop total rows,
        normalize.

        Returns the cleaned DataFrame ready for AI extraction. Deliberately
        does NOT call remove_numeric_rows() — that step blanks out every
        numeric-looking cell (including amounts), which was fine for the old
        label-matching pipeline this class originally served but would
        silently discard the amount on every line item for AI extraction,
        which needs the actual values, not just textual labels.
        """
        df, _ = self.read_sheet_with_openpyxl()
        df = self.filter_out_formula_rows()
        df = self.filter_out_total_rows(df)
        df = self.normalize_dataframe(df)
        return df


def to_extraction_grid(df: pd.DataFrame) -> list[list[str | None]]:
    """Convert a cleaned DataFrame (detect_structure()'s output) into a plain
    grid of cell text, stripped of the header row (column letters) and the
    row-number helper column — both are artifacts of the internal
    representation, not real sheet data, and would only confuse the AI
    extraction call or the fingerprint computed over this same grid."""
    body = df.iloc[1:, 1:]
    grid: list[list[str | None]] = []
    for _, row in body.iterrows():
        grid.append([None if v is None or v == "" or pd.isna(v) else str(v) for v in row])
    return grid
